from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk
import threading
import queue
import time


def idt_cust(val):
    return '=TEXT("%s","00000000000")'%val

def proper_string(val):
    return '=PROPER("%s")'%val

def lower_string(val):
    return '=LOWER("%s")'%val

def upper_string(val):
    return '=UPPER("%s")'%val
    
def phone(val):
    return '=TEXT("%s","(000) 000-0000")'%val
    
def sfdc_id(val):
    return '=TEXT("%s","000000000000000000")'%val

def zip_5(val):
    return '=TEXT("%s","00000")'%val

def initialize_val():
    #progress_txt.append(StringVar())
    for i in range(0,no_of_fields):
        accept.append(StringVar())
        check.append(StringVar())
        val.append(StringVar())
        val[i].set("")
        #print(val[i].get())

def get_val(*args):
    try:
        for i in range(0,no_of_fields):
            if (check[i].get()== '1'):
               val[i].set(accept[i].get())
            else:
                val[i].set("")
        
    except ValueError:
        return

def get_total_cells(val):
    entires_filled=0
    for i in range(0,no_of_fields):
        if val[i].get() != "":
            entires_filled += 1
            print(val[i].get())
    print(entires_filled)
    return entires_filled
            
        

def fill_val():
    print(a,b,no_cols_a,no_cols_b)
    for i in range(1,no_cols_a+1):
        for j in range(1,no_cols_b+1):
            if b_sheet.cell(row=1,column=j).value == a_sheet.cell(row=1,column=i).value:
                temp=a_sheet.cell(row=1,column=i).value
                #print(temp)
                for k in range(2,no_rows_a+1):
                    b_sheet.cell(row=k,column=j).value=formula[temp](a_sheet.cell(row=k,column=i).value)
                break
    b.save('output_cassandra.xlsx')

def fill_otherval_thread():
    global total_cells
    global cells_written
    cells_written=1
    total_cells=0
    total_cells=get_total_cells(val)*(no_rows_a-1)
    print("total_cells:"+str(total_cells))
    print("cells_written init:"+str(cells_written))
    for i in range(0,no_of_fields):
        #time.sleep(2)
        for j in range(1,no_cols_b+1):
            if b_sheet.cell(row=1,column=j).value == field_name[i]:
                temp=field_name[i]
                #print(temp)
                if val[i].get() != "":
                    for k in range(2,no_rows_a+1):
                        b_sheet.cell(row=k,column=j).value = val[i].get()
                        msg = "Progress "+(str(int((cells_written/total_cells)*100)))+"%"
                        queue.put(msg)
                        cells_written += 1
                        #print(cells_written)
                    print("Filled "+field_name[i]+" with value "+val[i].get())
    print("cells_written final:"+str(cells_written))
def periodiccall():
        checkqueue()
        if write_thread.is_alive():
            mainframe.after(1, periodiccall)
        else:
            print("Thread killed..")
            progress_txt.set('Writing')
            time.sleep(2)
			#modify path
            b.save('output_sample.xlsx')
            progress_txt.set('Process Complete')
            print("File written")
            submit_button.configure(state="active")
            cells_written=1
            total_cells=0

def checkqueue():
        #print('Q size: '+str(queue.qsize()))
        while queue.qsize():
            try:
                msg = queue.get(0)
                #self.listbox.insert('end', msg)
                #self.progressbar_write.title(msg)
                print("%complete.."+str(int(((cells_written-1)/total_cells)*100)))
                progressbar_write['value']=int(((cells_written-1)/total_cells)*100)
                progress_txt.set(str(int(((cells_written-1)/total_cells)*100))+"% complete")
                #progress_label.configure(textvariable=progress_txt.get())
            except queue.empty:
                print('Queue is empty..')
                pass

def spawnthread():
    submit_button.configure(state="disabled")
    
    global write_thread
  
    write_thread = threading.Thread(target=fill_otherval_thread)
    write_thread.start()
    periodiccall()

def fill_otherval():
    spawnthread()
    progressbar_write['value']=0

''' 
    progressbar_write.start()

    total_cells=get_total_cells(val)*(no_rows_a-1)
    print("total_cells:"+str(total_cells))
    cells_written=1

    for i in range(0,no_of_fields):
        for j in range(1,no_cols_b+1):
            if b_sheet.cell(row=1,column=j).value == field_name[i]:
                temp=field_name[i]
                #print(temp)
                if val[i].get() != "":
                    for k in range(2,no_rows_a+1):
                        b_sheet.cell(row=k,column=j).value = val[i].get()
                        progressbar_root.title("Progress "+(str(int((cells_written/total_cells)*100)))+"%")
                        progressbar_write.step(int((cells_written/total_cells)*100)) 
                        cells_written += 1
                        #print(cells_written)
                    print("Filled "+field_name[i]+" with value "+val[i].get())
            #break
    print(int(((cells_written-1)/total_cells)*100))
    if (int(((cells_written-1)/total_cells)*100)== 100):
        print("100% complete.")
        progressbar_write.stop()
    #progressbar_root.mainloop()
    b.save('output_cass_test.xlsx')
    print("File written")
'''
def OnFrameConfigure(event):
    canvas.configure(scrollregion=canvas.bbox("all"))

formula={'CAN':idt_cust,
         'Company':proper_string,
         'Phone':phone,
         'First Name':proper_string,
         'Last Name':proper_string,
         'Existing Account':sfdc_id,
         'Email':lower_string,
         'Street':proper_string,
         'City':proper_string,
         'State/Province':upper_string,
         'Zip/Postal Code':zip_5,
         'Channel':proper_string
         }

       
root = Tk()
root.title("SFDC Template")
root.grid_columnconfigure(0, weight=1)
root.grid_rowconfigure(0, weight=1)

canvas = Canvas(root,height=1000,width=1000)
canvas.grid(column=0, row=0,sticky=(N, W, E, S))

mainframe = ttk.Frame(canvas)
canvas.create_window((0,0), window=mainframe,anchor='nw')

vertical_sb = ttk.Scrollbar(root, orient=VERTICAL, command=canvas.yview)
canvas.configure(yscrollcommand=vertical_sb.set)
vertical_sb.grid(column=1,row=0,sticky=(N,S))


mainframe.bind("<Configure>", OnFrameConfigure)
#mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
#mainframe.columnconfigure(0, weight=1)
#mainframe.rowconfigure(0, weight=1)

queue = queue.Queue()

field_name = [
    "Accept Lead",
    "Campaign Id",
    "Status",
    "Primary Business Unit",
    "Lead Rating",
    "Lead Source",
    "Lead Source Detail",
    "Currency",
    "Country",
    "Solution Category",
    "Type of Solution",
    "Product Type",
    "Campaign Name",
    "Description",
     "Type of Solution",
    "Product Type",
    "Campaign Name",
    "Description",
]
no_of_fields=len(field_name)
field_name.sort()

#cells_written=1
#total_cells=0

accept = []
check = []
val = []
progress_txt=StringVar()
initialize_val()
for i in range(0,no_of_fields):
    #accept.append(StringVar())
    #check.append(StringVar())
    #val.append(StringVar())
    
    j=0
    ttk.Label(mainframe, text=field_name[i]).grid(column=j+1, row=i+2, sticky=W)
    ttk.Entry(mainframe, textvariable=accept[i], width=50).grid(column=j+2, row=i+2, sticky=W)
    ttk.Checkbutton(mainframe, variable=check[i], command=get_val).grid(column=j+3, row=i+2, sticky=W)
    ttk.Label(mainframe, text="Value Entered:").grid(column=j+4, row=i+2, sticky=W)
    ttk.Label(mainframe, textvariable=val[i]).grid(column=j+5, row=i+2, sticky=W)



submit_button=ttk.Button(mainframe, text="Submit", width=1, command=fill_otherval)
submit_button.grid(column=2, row=no_of_fields+2, sticky= (W, E))

ttk.Button(mainframe, text="Extract Fields from input file",width=50, command=fill_val).grid(column=2, row=1, sticky= (W, E))

progressbar_write = ttk.Progressbar(mainframe, mode='determinate', orient=HORIZONTAL, length=250)
progressbar_write.grid(column=2, row=no_of_fields+3, sticky= (W, E))
progressbar_write['maximum']=100


progress_label=ttk.Label(mainframe,textvariable=progress_txt)
progress_label.grid(column=3, row=no_of_fields+3, sticky= (W),columnspan=2)

for child in mainframe.winfo_children(): child.grid_configure(padx=50, pady=10)

#modify path
a=load_workbook('new.xlsx')
a_sheet=a.get_sheet_by_name('Sheet1')
no_cols_a=len(a_sheet.rows[1])
no_rows_a=len(a_sheet.rows)


#modify path
b=load_workbook('sfdc_format.xlsx')
b_sheet=b.active
no_cols_b=len(b_sheet.columns)
no_rows_b=len(b_sheet.rows)

root.mainloop()

